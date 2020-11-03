#! python3
# excel_to_csv.py - It will loop through a folder and work with the .xlsx files. The program will convert all the .xlsx files into .csv files.
# In case that the spreadsheet has multiples sheets, one .csv file must be created per sheet.
# the .csv filename must be <excelfilename>_<sheet_title>.csv

import openpyxl, csv, os

os.chdir(r'.\Chapter-14_csv_json_files\excel_files')
path = os.getcwd()

for excelFile in os.listdir(path):
    # Skip nom-xlsx files, load the workbook object.
    if not excelFile.endswith('.xlsx'):
        continue
    print(excelFile)
    wb = openpyxl.load_workbook(path + '\\' + excelFile)

    for sheetName in wb.get_sheet_names():
        # Loop through every sheet in the workbook.
        sheet = wb.get_sheet_by_name(sheetName)

        # Create the CSV filename from the Excel filename and sheet title.
        excelName = excelFile.split('.')[0]
        sheetName = sheet.title
        csvName = excelName + '_' + sheetName + '.csv'
        print(csvName)

        # Create the csv.writer object for this CSV file.
        csvFileObj = open(csvName, 'w')
        csvFile = csv.writer(csvFileObj)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []    # Append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                myData = sheet.cell(rowNum, colNum).value
                rowData.append(myData)
            # Write the rowData list to the CSV file.
            csvFile.writerow(rowData)
        
        csvFileObj.close()