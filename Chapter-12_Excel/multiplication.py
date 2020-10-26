#! python3
# multiplication.py - Creates a excel spreadsheet with a 
# multiplication table inside


import openpyxl
from openpyxl.styles import Alignment
from os import chdir

# set up directory
chdir(r'Chapter-12_Excel')

# create a workbook and a sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# create a function to build the table
def createResults(size):
    # size: integer
    # return: list
    myTable = []

    for colnum in range(1, size+1):
        # myTable will have a list inside for each row
        dataTable = []
        for rownum in range(1, size+1):
            result = colnum * rownum
            dataTable.append(result)
        myTable.append(dataTable)
    
    # append the data to the sheet
    for row in myTable:
        sheet.append(row)
    
    createHeader(size)


# create headers in Excel
def createHeader(size):
    sheet.insert_cols(1)
    sheet.insert_rows(1)
    for i in range(1, size+1):
        sheet.cell(1, i+1).value = i    # column
        sheet.cell(1, i+1).alignment = Alignment(horizontal='center', textRotation=15)
        sheet.cell(1, i+1).font = openpyxl.styles.Font(bold=True)

        sheet.cell(i+1, 1).value = i    # row
        sheet.cell(i+1, 1).alignment = Alignment(horizontal='center', textRotation=15)
        sheet.cell(i+1, 1).font = openpyxl.styles.Font(bold=True)
# send it to the worksheet
createResults(50)

# save

workbook.save(filename='multiplicationtable.xlsx')
