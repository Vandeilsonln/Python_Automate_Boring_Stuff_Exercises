import openpyxl
import os
import json


os.chdir(r'.\Chapter-12_Excel')

myWorkbook = openpyxl.load_workbook('example.xlsx')

mySheet = myWorkbook.active

print(mySheet['A1:C2'])

print('-'*40)

for row in mySheet.iter_rows(min_row=1, max_row=4, min_col=1, max_col=5):
    print(row)

print('-'*40)

for column in mySheet.iter_cols(min_row=1, max_row=4, min_col=1, max_col=5):
    print(column)

print('-'*40)

# for column in mySheet.columns:
#    print(column)

print('-'*40)
print('-'*40)
print('-'*40)

for value in mySheet.iter_rows(min_row=1, max_row=1, values_only=True):
    print(value)
print('-'*40)
for value in mySheet.iter_rows(min_row=2, max_row=5, min_col=3, max_col=9, values_only=True):
    print(value)

print('-'*40)
print('-'*40)
print('-'*40)

databaseCompanies = {}
for i in mySheet.iter_rows(min_col=3, max_col=9, min_row=2, values_only=True):
    companyName = i[0]
    companyInfo = {
        'address': i[1],
        'phone No': i[4],
        'e-mail': i[5],
        'website': i[6]
    }
    databaseCompanies[companyName] = companyInfo

# Using JSON here to be able to format the output for displaying later
print(json.dumps(databaseCompanies))

print('-'*40)

