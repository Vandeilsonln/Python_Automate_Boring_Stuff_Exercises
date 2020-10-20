from datetime import datetime
from openpyxl import load_workbook
from classes import company
import os
from mapping import COMPANY_NAME, COMPANY_ADDRESS, COMPANY_PHONE, COMPANY_EMAIL, COMPANY_WEB


# Using the read_only method since you're not gonna be editing the spreadsheet.
os.chdir(r'.\Chapter-12_Excel')

myWorkbook = load_workbook(filename='example.xlsx', read_only=True)
sheet = myWorkbook.active

dbCompanies = []

# Using the values_only because you just want to return the cell value
for row in sheet.iter_rows(min_row=2, values_only=True):
    dbCompany = company(name=row[COMPANY_NAME],
    address=row[COMPANY_ADDRESS],
    phone=row[COMPANY_PHONE],
    email=row[COMPANY_EMAIL],
    web=row[COMPANY_WEB]
    )
    dbCompanies.append(dbCompany)

for i in dbCompanies:
    print(i)