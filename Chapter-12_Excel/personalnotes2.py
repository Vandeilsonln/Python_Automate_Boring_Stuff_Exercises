from openpyxl import load_workbook
from os import chdir


chdir(r'.\Chapter-12_Excel')

def print_rows(CurSheet):
    for i in CurSheet.iter_rows(values_only=True):
        print(i)

workbook = load_workbook(filename='hello_world.xlsx')

sheet = workbook.active

# cell = sheet['A1']
# cell.value = 'Hey'
# print(cell.value)

# print_rows(sheet)

# sheet['B10'] = 'test'
# print_rows(sheet)

# # Insert a column BEFORE the existing column 1 ('A')
# sheet.insert_cols(idx=1)
# print_rows(sheet)

# # Insert 5 columns Between column 2 ('B') and 3 ('C')
# sheet.insert_cols(idx=3, amount=5)
# print_rows(sheet)

# # delete th created columns
# sheet.delete_cols(idx=3, amount=5)
# sheet.delete_cols(idx=1)
# print_rows(sheet)

# sheet.delete_rows(idx=3, amount=7)
# print_rows(sheet)

# print(workbook.sheetnames)

# sheet.title = 'Python_Test'

# print(workbook.sheetnames)

# testSheet = workbook.create_sheet('Testing_openpyxl')

# print(workbook.sheetnames)

# idxSheet = workbook.create_sheet('Indexed', 0)

# print(workbook.sheetnames)

# workbook.remove(idxSheet)
# workbook.remove(testSheet)

# print(workbook.sheetnames)

# workbook.copy_worksheet(sheet)

# print(workbook.sheetnames)
# workbook.save('hello_world.xlsx')

sheet.freeze_panes = 'C2'
workbook.save('hello_world.xlsx')