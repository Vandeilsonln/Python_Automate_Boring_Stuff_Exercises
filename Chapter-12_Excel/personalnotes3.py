#! python3

import openpyxl
from os import chdir

chdir(r'Chapter-12_Excel')

workbook = openpyxl.load_workbook('personalnotes.xlsx')
sheet = workbook.get_sheet_by_name('Personal notes 3')

print(sheet.max_column)
print(sheet.max_row)
print(sheet.dimensions)

max_column1 = sheet.dimensions.split(':')[1][-2::1]
print(max_column1)
# So, when you want to append new data into the sheet, start from that value+1